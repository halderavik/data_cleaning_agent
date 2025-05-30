import os
import shutil
from fastapi import APIRouter, UploadFile, File, HTTPException, status, Depends, Form
from fastapi.responses import JSONResponse
from tempfile import NamedTemporaryFile
from typing import Any
from pandas import read_csv
import openpyxl
import numpy as np
import logging
import traceback
from sqlalchemy.orm import Session
from uuid import UUID
import math

from backend.security import get_current_active_user
from backend.models import User, DataFile, Project
from backend.services.spss_processor import SPSSProcessor
from backend.cleaning_engine import CleaningEngine
from backend.database.base import get_db
from sqlalchemy.exc import NoResultFound

router = APIRouter(
    prefix="/files",
    tags=["files"],
    responses={404: {"description": "Not found"}},
)

def to_serializable(obj):
    if isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, np.floating):
        if np.isnan(obj) or np.isinf(obj):
            return None
        return float(obj)
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, dict):
        return {k: to_serializable(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [to_serializable(i) for i in obj]
    return obj

@router.post("/analyze-spss", summary="Upload and analyze SPSS file")
async def analyze_spss_file(
    file: UploadFile = File(...)
) -> Any:
    """
    Upload an SPSS (.sav) file and return its structure analysis.
    Args:
        file (UploadFile): The uploaded SPSS file.
    Returns:
        dict: Structure analysis of the SPSS file.
    Raises:
        HTTPException: If file is not a valid SPSS file or analysis fails.
    """
    if not file.filename.lower().endswith(".sav"):
        raise HTTPException(status_code=400, detail="Only SPSS (.sav) files are supported.")

    try:
        with NamedTemporaryFile(delete=False, suffix=".sav") as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name
        processor = SPSSProcessor()
        structure = processor.load_file(tmp_path)
        # Load data as DataFrame for cleaning
        import pyreadstat
        df, meta = pyreadstat.read_sav(tmp_path)
        cleaning_engine = CleaningEngine()
        cleaning_results = cleaning_engine.process_data(df)
    except Exception as e:
        logging.error(f"Failed to process SPSS file: {str(e)}")
        logging.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Failed to process SPSS file: {str(e)}")
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
    return JSONResponse(content=to_serializable({"structure": structure, "cleaning_results": cleaning_results}))

@router.post("/analyze-csv", summary="Upload and analyze CSV file")
async def analyze_csv_file(
    file: UploadFile = File(...)
) -> Any:
    """
    Upload a CSV file and return its structure analysis.
    """
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only CSV files are supported.")
    try:
        with NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name
        df = read_csv(tmp_path)
        structure = {
            "columns": [to_serializable(col) for col in df.columns.tolist()],
            "row_count": int(len(df)),
            "preview": [{k: to_serializable(v) for k, v in row.items()} for row in df.head(5).to_dict(orient="records")]
        }
        cleaning_engine = CleaningEngine()
        cleaning_results = cleaning_engine.process_data(df)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process CSV file: {str(e)}")
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
    return JSONResponse(content=to_serializable({"structure": structure, "cleaning_results": cleaning_results}))

@router.post("/analyze-excel", summary="Upload and analyze Excel file")
async def analyze_excel_file(
    file: UploadFile = File(...)
) -> Any:
    """
    Upload an Excel file (.xls, .xlsx) and return its structure analysis.
    """
    if not (file.filename.lower().endswith(".xls") or file.filename.lower().endswith(".xlsx")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xls, .xlsx) are supported.")
    try:
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        sheet = wb.active
        columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        row_count = sheet.max_row - 1
        preview = []
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=6), start=1):
            preview.append({columns[j]: cell.value for j, cell in enumerate(row)})
        import pandas as pd
        data = list(sheet.values)
        df = pd.DataFrame(data[1:], columns=data[0])
        structure = {
            "columns": [to_serializable(col) for col in columns],
            "row_count": int(row_count),
            "preview": [
                {k: to_serializable(v) for k, v in row.items()} for row in preview
            ]
        }
        cleaning_engine = CleaningEngine()
        cleaning_results = cleaning_engine.process_data(df)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process Excel file: {str(e)}")
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass
    return JSONResponse(content=to_serializable({"structure": structure, "cleaning_results": cleaning_results}))

@router.post("/upload-and-register", summary="Upload file and register in DB")
async def upload_and_register_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Upload a file, save to disk, and register in DataFile DB table.
    Returns file_id for analysis endpoints.
    No project_id is required or set.
    Also runs basic cleaning and returns structure and cleaning results.
    """
    import traceback
    import openpyxl
    import pandas as pd
    import pyreadstat
    from backend.cleaning_engine import CleaningEngine
    try:
        # Save file to disk
        ext = os.path.splitext(file.filename)[1].lower()
        file_type = (
            'sav' if ext == '.sav' else
            'csv' if ext == '.csv' else
            'excel' if ext in ['.xls', '.xlsx'] else
            'unknown'
        )
        if file_type == 'unknown':
            raise HTTPException(status_code=400, detail="Unsupported file type.")
        uploads_dir = os.path.join(os.getcwd(), 'uploads')
        os.makedirs(uploads_dir, exist_ok=True)
        file_path = os.path.join(uploads_dir, file.filename)
        with open(file_path, 'wb') as out_file:
            shutil.copyfileobj(file.file, out_file)
        file_size = os.path.getsize(file_path)
        # Register in DB without project_id
        data_file = DataFile(
            original_filename=file.filename,
            file_size=file_size,
            file_type=file_type,
            upload_status='completed'
        )
        db.add(data_file)
        db.commit()
        db.refresh(data_file)
        # Run basic cleaning and structure analysis
        structure = None
        cleaning_results = None
        check_docs = None
        try:
            if file_type == 'sav':
                df, meta = pyreadstat.read_sav(file_path)
                structure = {
                    "columns": [str(col) for col in df.columns.tolist()],
                    "row_count": int(len(df)),
                    "preview": df.head(5).to_dict(orient="records")
                }
            elif file_type == 'csv':
                df = pd.read_csv(file_path)
                structure = {
                    "columns": [str(col) for col in df.columns.tolist()],
                    "row_count": int(len(df)),
                    "preview": df.head(5).to_dict(orient="records")
                }
            elif file_type == 'excel':
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet = wb.active
                columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                row_count = sheet.max_row - 1
                preview = []
                for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=6), start=1):
                    preview.append({columns[j]: cell.value for j, cell in enumerate(row)})
                import pandas as pd
                data = list(sheet.values)
                df = pd.DataFrame(data[1:], columns=data[0])
                structure = {
                    "columns": [str(col) for col in columns],
                    "row_count": int(row_count),
                    "preview": preview
                }
            # Run cleaning
            cleaning_engine = CleaningEngine()
            cleaning_results = cleaning_engine.process_data(df)
            check_docs = cleaning_engine.get_check_documentation()
        except Exception as clean_exc:
            cleaning_results = {"error": str(clean_exc)}
            check_docs = {}
        return to_serializable({
            "file_id": str(data_file.id),
            "structure": structure,
            "cleaning_results": cleaning_results,
            "check_docs": check_docs
        })
    except Exception as e:
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))

@router.post('/create-project', summary='Create a new project and return its ID')
async def create_project(
    name: str = Form(...),
    description: str = Form(None),
    db: Session = Depends(get_db)
):
    print(f"[create-project] Received name: {name!r}, description: {description!r}")
    try:
        user = db.query(User).first()
        if not user:
            print("[create-project] No user found in DB.")
            raise HTTPException(status_code=400, detail='No user found to assign as project owner.')
        project = Project(
            name=name,
            description=description,
            status='active',
            owner_id=user.id
        )
        db.add(project)
        db.commit()
        db.refresh(project)
        print(f"[create-project] Project created with id: {project.id}")
        return {'project_id': str(project.id)}
    except Exception as e:
        import traceback
        print(f"[create-project] Exception: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e)) 